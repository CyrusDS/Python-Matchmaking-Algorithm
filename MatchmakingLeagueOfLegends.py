import openpyxl
from pathlib import Path
import matplotlib.pyplot as plt
import numpy as np

def expectProb(A, B):
    #Determines the expected chance to win for Team A, against Team B
    #A is a rating for Team A
    #B is a rating for Team B

    return 1 / (1 + 10 ** ((B - A) / 400))


def elo(old, exp, score, k=32):
    #Determine updated Elo rating for a team 
    #old: Team's current Elo rating
    #exp: expected chance to win for the Team
    #1 = won, 0 = loss
    #K-factor for algorithm, default is 32
    
    return old + k * (score - exp)

class Team:
    def __init__(self, name, rating, wins, losses):
        self.name = name
        self.rating = rating
        self.wins = wins
        self.losses = losses
        
Init = Team("MEAN RATING", 1500, 0, 0)#set up all team names, initial rating, wins and losses. 1500 is the default mean rating for this relative scale.
HLE = Team("Hanwha", 1500, 0, 0)
PCE = Team("PEACE", 1500, 0, 0)
RED = Team("RED Canids", 1500, 0, 0)
INF = Team("Infinity", 1500, 0, 0)
DFM = Team("DetonatioN FocusMe", 1500, 0, 0)
C9 = Team("CloudNine", 1500, 0, 0)
GALA = Team("Galatasaray", 1500, 0, 0)
BYG = Team("Beyond", 1500, 0, 0)
UOL = Team("Unicorns of Love", 1500, 0, 0)
DWG = Team("DWG KIA", 1500, 0, 0)
RGE = Team("Rogue", 1500, 0, 0)
FPX = Team("FunPlus Phoenix", 1500, 0, 0)
T1 = Team("T1", 1500, 0, 0)
EDG = Team("EDG", 1500, 0, 0)
RNG = Team("RNG", 1500, 0, 0)
PSG = Team("PSG Talon", 1500, 0, 0)
FNC = Team("Fnatic", 1500, 0, 0)
GenG = Team("Gen.G", 1500, 0, 0)
MAD = Team("MAD Lions", 1500, 0, 0)
LNG = Team("LNG", 1500, 0, 0)
TL = Team("Team Liquid", 1500, 0, 0)
HThiev = Team("100 Thieves", 1500, 0, 0)

TeamsList = [Init,HLE,PCE,RED,INF,DFM,C9,GALA,BYG,UOL,DWG,RGE,FPX,T1,EDG,RNG,PSG,FNC,GenG, MAD, LNG, TL, HThiev] #put all teams into a list


xlsx_file = Path(Path.home(), 'Documents', 'test', 'datatest.xlsx') #import from documents, test folder, datatest file
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active

print("Total Matches Analyzed: ",sheet.max_row) #indicate data analyzed

for i in range(1, sheet.max_row+1): #iterate each row
    FirstTeam = TeamsList[0] #init each team as default
    SecondTeam = TeamsList[0]#init each team as default
    
    for j in range(1, sheet.max_column+1): #iterate through the column from j = 1 to j = 4
        cell_obj = sheet.cell(row=i, column=j)
        if(j == 1):
            for f in range(len(TeamsList)):
                #print("Checking1: ", TeamsList[f].name, cell_obj.value)
                if(TeamsList[f].name == cell_obj.value):
                    FirstTeam = TeamsList[f]
                    break
        if(j == 2):
            Team1Score = cell_obj.value
        if(j == 4):
           for d in range(len(TeamsList)):
                #print("Checking2: ", TeamsList[d].name, cell_obj.value)
                if(TeamsList[d].name == cell_obj.value):
                    SecondTeam = TeamsList[d]
                    break
           if(Team1Score == 1): #Blue won.
                #print(FirstTeam.name, "vs" , SecondTeam.name, "Winner: ", FirstTeam.name, "r: ", i, "j: ", j)

                FirstTeam.rating = elo(FirstTeam.rating, expectProb(FirstTeam.rating, SecondTeam.rating), 1)
                SecondTeam.rating = elo(SecondTeam.rating, expectProb(SecondTeam.rating, FirstTeam.rating), 0)
                FirstTeam.wins +=1
                SecondTeam.losses +=1
           else: #Red won
                #print(FirstTeam.name, "vs" , SecondTeam.name, "Winner: ", SecondTeam.name, "r: ", i, "j: ", j)
                FirstTeam.rating = elo(FirstTeam.rating, expectProb(FirstTeam.rating, SecondTeam.rating), 0)
                SecondTeam.rating = elo(SecondTeam.rating, expectProb(SecondTeam.rating, FirstTeam.rating), 1)
                FirstTeam.losses +=1
                SecondTeam.wins +=1

TeamsList.sort(key=lambda x: x.rating, reverse=True) #sort list by ascending rating
x, y = [], []
for i in range(len(TeamsList)): #print entire list showing team name, rating, wins, losses
    print(TeamsList[i].name, "{0:.2f}".format(TeamsList[i].rating), "W:" ,TeamsList[i].wins, "L:", TeamsList[i].losses)
    x.append(TeamsList[i].wins - TeamsList[i].losses) #unused, ignore
    y.append(TeamsList[i].rating)




count, bins_count = np.histogram(y,bins=10)
pdf = count / sum(count)
cdf = np.cumsum(pdf)

plt.plot(bins_count[1:],cdf,label="CDF for 22 League of Legends Teams")
plt.legend()
plt.xlabel("Skill Rating")
plt.ylabel("Cumulative Percentage")
plt.show()
